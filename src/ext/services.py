import os
from fastapi import UploadFile, File
from openpyxl import load_workbook
from sqlalchemy import insert

from src import cfg
from src.database import get_async_session, async_session_maker
from src.models import EXT_M


async def ext_get_all_count():
    content = {"msg": f"Unknown error"}
    try:
        async with get_async_session() as session:
            all_count = await session.query(EXT_M).count()
            content = {"msg": "Success", "count": all_count}
            session.close()
            return content
    except Exception as e:
        cont_err = f"fail. can't read count from table ({EXT_M.__tablename__})"
        content = {"msg": "error", "data": f"Exception occurred {str(e)} . {cont_err}"}
        print(content)
    finally:
        pass
        # if session is not None:
        #     session.close()
    return content


async def ext_upload_file(file: UploadFile = File(...)):
    content = {"msg": f"Unknown error"}
    try:

        if not os.path.isdir(cfg.FOLDER_UPLOAD):
            os.mkdir(cfg.FOLDER_UPLOAD)
        # file_in = file.filename
        file_out = file.filename.replace(" ", "-")

        file_name = os.path.join(os.getcwd(), cfg.FOLDER_UPLOAD, file_out)

        if os.path.exists(file_name):
            os.remove(file_name)
            print(f"{file_name} - deleted !")

        contents = file.file.read()
        with open(file_name, 'wb') as f:
            f.write(contents)

        all_count = 1
        content = {"msg": "Success", "count": all_count, "filename": file.filename}

        await excel2db(file_name)

        return content
    except Exception as e:
        str_err = "Exception occurred " + str(e)
        content = {"msg": "error", "data": f"Exception occurred {str(e)} . {str_err}"}
        print(str_err)
    finally:
        file.file.close()

    return content


async def excel2db(file_in: str):
    print(f"File input {file_in}")

    # await ADOC_M.objects.delete(each=True)
    # await AUTHOR_M.objects.delete(each=True)

    # async with get_async_session() as session:
    # all_count = await session.query(EXT_M).count()
    # content = {"msg": "Success", "count": all_count}
    # session.close()
    async with async_session_maker() as session:
        # Define variable to load the wookbook
        # wookbook = openpyxl.load_workbook(file_in)

        # session.query(EXT_M).delete()
        stmt = f"TRUNCATE {EXT_M.__tablename__} RESTART IDENTITY;"
        print(stmt)
        result = await session.execute(stmt)
        await session.commit()


        wookbook = load_workbook(file_in)

        # Define variable to read the active sheet:
        worksheet = wookbook.active

        # Iterate the loop to read the cell values
        # worksheet.max_row
        # worksheet.max_column

        cnt = 0
        try:

            for value in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=5,
                                             values_only=True):
                # print(f"{cnt}")
                cnt = cnt + 1
                str_ext = ''
                str_cat = ''
                str_desc = ''
                str_prod = ''
                str_is_project = ''
                is_project = False

                if value[0]:  # Ext
                    str_ext = str(value[0]).strip()

                if value[1]:  # Category
                    str_cat = str(value[1]).strip()

                if value[2]:  # Description
                    str_desc = str(value[2]).strip()

                if value[3]:  # Product
                    str_prod = str(value[3]).strip()

                if value[4]:  # is_project
                    str_is_project = str(value[4]).strip()
                    # if str_is_project.isdigit():
                    #     if int(str_is_project) == 1:
                    #         is_project = True

                print(f"{str_ext} {str_cat} {str_desc} {str_prod} {str_is_project} ")

                stmt = insert(EXT_M).values(ext=str_ext, category=str_cat, description=str_desc, product=str_prod,
                                            is_project=str_is_project)
                print(stmt)
                result = await session.execute(stmt)
                await session.commit()

                # new_src = FILE_SRC_M(folder_src=folder)
                # session.add(new_src)
                # await session.commit()

                # conn.commit()
                # if value[14]:  # Отчет
                #     str_tmp1 = str(value[14]).strip().replace("_x001E_", "")
                #     report_name = str_cleanup(str_tmp1)

                # ).save()
            await session.close()

        except Exception as e:
            content = {"msg": "error", "data": f"Exception occurred {str(e)}"}
            print(content)
