import os
from fastapi import UploadFile, File
from openpyxl import load_workbook
from sqlalchemy import insert, func, select, text

from sqlalchemy import distinct

from src import cfg
from src.db.db import async_session_maker
from src.models import M_EXT


# from src.database import get_async_session, async_session_maker, engine
# from src.models import EXT_M


# from src.database import async_session


# @async_session
# async def async_fetch_categories(async_session):
#     content = {"msg": f"Unknown error"}
#     stmt = select(EXT_M).filter()
#     result = await async_session.execute(stmt)
#     return result.scalars().all()


async def ext_get_all_count():
    content = {"msg": f"error"}
    try:
        async with async_session_maker() as session:
            res = await session.scalar(select(func.count(M_EXT.id)))
            content = {"msg": "Success", "count": res}
            return content
    except Exception as e:
        cont_err = f"fail. can't read ext from table ({M_EXT.__tablename__})"
        content = {"msg": "error", "data": f"Exception occurred {str(e)} . {cont_err}"}
        print(content)
    finally:
        if session is not None:
            await session.close()
    return content


async def ext_get_all():
    content = {"msg": f"Unknown error"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_EXT)
                .order_by(M_EXT.ext)
            )
            _all = res.all()
            cnt = len(_all)
            content = {"msg": "Success", "count": cnt, "data": _all}
            return content
    except Exception as e:
        cont_err = f"fail. can't read ext from table ({M_EXT.__tablename__})"
        content = {"msg": "error", "data": f"Exception occurred {str(e)} . {cont_err}"}
        print(content)
    finally:
        if session is not None:
            await session.close()
    return content


async def ext_get_uniq():
    content = {"msg": f"Unknown error"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_EXT)
                .distinct()
                .order_by(M_EXT.ext)
            )
            _all = res.all()
            cnt = len(_all)
            content = {"msg": "Success", "count": cnt, "data": _all}
            return content
    except Exception as e:
        cont_err = f"fail. can't read ext from table ({M_EXT.__tablename__})"
        content = {"msg": "error", "data": f"Exception occurred {str(e)} . {cont_err}"}
        print(content)
    finally:
        if session is not None:
            await session.close()
    return content


async def ext_get_info_by_ext(ext: str):
    content = {"msg": f"Unknown error"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_EXT)
                .where(M_EXT.ext == ext)
            )
            _all = res.all()
            cnt = len(_all)
            content = {"msg": "Success", "count": cnt, "data": _all}
            return content
    except Exception as e:
        cont_err = f"fail. can't read ext from table ({M_EXT.__tablename__})"
        content = {"msg": "error", "data": f"Exception occurred {str(e)} . {cont_err}"}
        print(content)
    finally:
        if session is not None:
            await session.close()
    return content


async def ext_get_ext_by_cat(cat: str):
    content = {"msg": f"Unknown error"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_EXT)
                .where(M_EXT.category == cat)
            )
            _all = res.all()
            cnt = len(_all)
            content = {"msg": "Success", "count": cnt, "data": _all}
            return content
    except Exception as e:
        cont_err = f"fail. can't read ext from table ({M_EXT.__tablename__})"
        content = {"msg": "error", "data": f"Exception occurred {str(e)} . {cont_err}"}
        print(content)
    finally:
        if session is not None:
            await session.close()
    return content


async def ext_get_ext_by_product(product: str):
    content = {"msg": f"Unknown error"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_EXT)
                .where(M_EXT.product == product)
            )
            _all = res.all()
            cnt = len(_all)
            content = {"msg": "Success", "count": cnt, "data": _all}
            return content
    except Exception as e:
        cont_err = f"fail. can't read ext from table ({M_EXT.__tablename__})"
        content = {"msg": "error", "data": f"Exception occurred {str(e)} . {cont_err}"}
        print(content)
    finally:
        if session is not None:
            await session.close()
    return content


async def ext_get_ext_by_project():
    content = {"msg": f"Unknown error"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_EXT)
                .where(M_EXT.is_project == '1')
            )
            _all = res.all()
            cnt = len(_all)
            content = {"msg": "Success", "count": cnt, "data": _all}
            return content
    except Exception as e:
        cont_err = f"fail. can't read ext from table ({M_EXT.__tablename__})"
        content = {"msg": "error", "data": f"Exception occurred {str(e)} . {cont_err}"}
        print(content)
    finally:
        if session is not None:
            await session.close()
    return content


async def ext_upload_file(file: UploadFile = File(...)):
    content = {"msg": f"Unknown error"}
    try:

        if not os.path.isdir(cfg.FOLDER_UPLOAD):
            os.mkdir(cfg.FOLDER_UPLOAD)
        # file_in = file.filename
        file_out = file.filename.replace(" ", "-")

        file_name = os.path.join(os.getcwd(), cfg.FOLDER_UPLOAD, file_out)

        if os.path.exists(file_name):
            os.remove(file_name)
            print(f"{file_name} - deleted !")

        contents = file.file.read()
        with open(file_name, 'wb') as f:
            f.write(contents)

        all_count = 1
        content = {"msg": "Success", "count": all_count, "filename": file.filename}

        await excel2db(file_name)

        return content
    except Exception as e:
        str_err = "Exception occurred " + str(e)
        content = {"msg": "error", "data": f"Exception occurred {str(e)} . {str_err}"}
        print(str_err)
    finally:
        file.file.close()

    return content


async def excel2db(file_in: str):
    print(f"File input {file_in}")

    content = {"msg": "Success", "count": 0}

    async with async_session_maker() as session:

        # session.query(EXT_M).delete()
        stmt = f"TRUNCATE {M_EXT.__tablename__} RESTART IDENTITY;"
        print(stmt)
        result = await session.execute(text(stmt))
        await session.commit()

        _workbook = load_workbook(file_in)

        # Define variable to read the active sheet:
        worksheet = _workbook.active

        # Iterate the loop to read the cell values
        # worksheet.max_row
        # worksheet.max_column

        cnt = 0
        try:

            for value in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=5,
                                             values_only=True):
                # print(f"{cnt}")
                cnt = cnt + 1
                str_ext = ''
                str_cat = ''
                str_desc = ''
                str_prod = ''
                str_is_project = ''
                is_project = False

                if value[0]:  # Ext
                    str_ext = str(value[0]).strip()

                if value[1]:  # Category
                    str_cat = str(value[1]).strip()

                if value[2]:  # Description
                    str_desc = str(value[2]).strip()

                if value[3]:  # Product
                    str_prod = str(value[3]).strip()

                if value[4]:  # is_project
                    str_is_project = str(value[4]).strip()
                    # if str_is_project.isdigit():
                    #     if int(str_is_project) == 1:
                    #         is_project = True

                print(f"{str_ext} {str_cat} {str_desc} {str_prod} {str_is_project} ")

                stmt = insert(M_EXT).values(ext=str_ext, category=str_cat, description=str_desc, product=str_prod,
                                            is_project=str_is_project)
                # conn.execute(some_table.insert().values(foo="bar"))
                print(stmt)
                await session.execute(stmt)
                await session.commit()
            await session.close()

        except Exception as e:
            content = {"msg": "error", "data": f"Exception occurred {str(e)}"}
            print(content)
