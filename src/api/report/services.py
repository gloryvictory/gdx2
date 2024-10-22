import os
import re
import shutil
from datetime import datetime

import openpyxl
import sqlalchemy
from fastapi import UploadFile, File
from sqlalchemy import text, insert, select, func
from sqlalchemy.orm import mapped_column
from starlette.requests import Request

from src import cfg
from src.api.report.utils import str_get_folder, str_get_full_path_with_format, str_tgf_format, str_clean, \
    str_get_last_folder, str_get_folder_src, str_get_rgf, list_str_clean
from src.db.db import async_session_maker
from src.models import M_REPORT_TGF, M_R_AUTHOR, M_R_LIST, M_R_SUBRF, M_R_ORG, M_R_AREA, M_R_FIELD, M_R_LU, M_R_PI, \
    M_R_VID_RAB, M_HISTORY, M_R_MESSAGE
from src.schemas import S_R_MESSAGE, S_R_MESSAGE_POST
from src.utils.mystrings import str_cleanup
from src.api.celery.tasks import report_update_author, report_update_subrf, report_update_org, report_update_area, \
    report_update_list, report_update_field, report_update_lu, report_update_pi, report_update_vid_rab, \
    report_index_create_task, report_update_history


async def report_upload_file(file: UploadFile = File(...)):
    content = {"msg": f"Unknown error"}
    try:
        if not os.path.isdir(cfg.FOLDER_UPLOAD):
            os.mkdir(cfg.FOLDER_UPLOAD)
        file_in = file.filename
        file_out = file.filename.replace(" ", "-")

        file_name = os.path.join(os.getcwd(), cfg.FOLDER_UPLOAD, file_out)

        if os.path.exists(file_name):
            os.remove(file_name)
            print(f"{file_name} - deleted !")

        contents = file.file.read()
        with open(file_name, 'wb') as f:
            f.write(contents)

        # adoc_hist = ADOC_HISTORY_M(file_in=file_in, file_out=file_out, file_out_path=file_name)
        # await adoc_hist.upsert()

        all_count = 1
        content = {"msg": "Success", "count": all_count, "filename": file.filename}

        # await adoc_excel_file_read(file_name)
        return content
    except Exception as e:
        str_err = "Exception occurred " + str(e)
        content = {"msg": f"Error. There was an error uploading the file", "err": str(e)}
        print(str_err)
        # log.info(str_err)
    finally:
        file.file.close()

    return content


#
async def report_update():
    content = {"msg": "Fail"}
    try:
        content = {"msg": "Success", "count": 1}
        file_report_in = os.path.join(cfg.FOLDER_BASE, cfg.FOLDER_UPLOAD, cfg.FILE_REPORT_NAME)
        await report_excel_file_read(str(file_report_in))
        return content
    except Exception as e:
        content = {"msg": f"reload fail. can't... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


# обновляем через Celery
async def report_update_from_file_with_task():
    content = {"msg": "Fail"}
    try:
        time1 = datetime.now()
        file_report_in = os.path.join(cfg.FOLDER_REPORT, cfg.FILE_REPORT_NAME)
        file_report_out = os.path.join(cfg.FOLDER_BASE, cfg.FOLDER_UPLOAD, cfg.FILE_REPORT_NAME)
        shutil.copyfile(file_report_in, file_report_out)
        print(f"{file_report_in} is copied to {file_report_out} - OK.")
        await report_excel_file_read(str(file_report_out))
        report_update_author.delay()
        report_update_subrf.delay()
        report_update_org.delay()
        report_update_area.delay()
        report_update_list.delay()
        report_update_field.delay()
        report_update_lu.delay()
        report_update_pi.delay()
        report_update_vid_rab.delay()
        report_index_create_task.delay()

        time2 = datetime.now()
        print(f"Обработано: {file_report_out}. Total time:  + {str(time2 - time1)}")
        content = {"msg": "Success", "count": 1}
        return content
    except Exception as e:
        content = {"msg": f"reload fail. can't... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_excel_file_read(file_in: str):
    print(f"File input {file_in}")
    # Define variable to load the wookbook
    wookbook = openpyxl.load_workbook(file_in)

    # Define variable to read the active sheet:
    worksheet = wookbook.active

    # Iterate the loop to read the cell values
    # worksheet.max_row
    # worksheet.max_column
    # 18 колонок
    cnt = 0
    min_row = 2
    max_row = worksheet.max_row
    min_col = 1
    max_col = 30
    engine = sqlalchemy.create_engine(cfg.DB_DSN2)

    try:
        authors = []
        # create session and add objects
        async with async_session_maker() as session:
            # truncate table
            stmt = text(f"TRUNCATE {M_REPORT_TGF.__tablename__} RESTART IDENTITY;")
            await session.execute(stmt)
            await session.commit()
            for value in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col,
                                             values_only=True):
                print(f"{cnt}")
                cnt = cnt + 1
                folder_root = ''
                folder_link = ''
                folder_short = ''
                folder_name = ''
                rgf = ''
                tgf_hmao = ''
                tgf_ynao = ''
                tgf_kras = ''
                tgf_ekat = ''
                tgf_omsk = ''
                tgf_novo = ''
                tgf_tomsk = ''
                tgf_more = ''
                tgf_tmn = ''
                tgf_kurgan = ''
                tgf = ''
                report_name = ''
                author_name = ''
                year_str = ''
                year_int = 0
                territory_name = ''
                subrf_name = ''
                list_name = ''
                part_name = ''
                areaoil = ''
                field = ''
                lu = ''
                pi_name = ''
                fin_name = ''
                org_name = ''
                zsniigg_report = ''
                inf_report = ''
                vid_rab = ''

                comments = ''

                if value[0]:  # Путь полный
                    folder_root = str_get_full_path_with_format(str(value[0]))
                    folder_link = str_get_folder_src(folder_root)  # Гиперссылка
                    folder_short = str_get_folder(folder_root)
                    folder_name = str_get_last_folder(folder_root)
                    print(folder_root)

                if value[1]:  # Инв. номер РГФ
                    tmp_str = str(value[1])
                    if tmp_str.isdecimal():
                        rgf = tmp_str
                    else:
                        rgf = ''
                    print(rgf)
                if value[2]:
                    tgf_hmao = str_tgf_format(str(value[2]))
                    # print(tgf_hmao)
                if value[3]:
                    tgf_ynao = str_tgf_format(str(value[3]))
                    # print(tgf_ynao)
                if value[4]:
                    tgf_kras = str_tgf_format(str(value[4]))
                    # print(tgf_kras)
                if value[5]:
                    tgf_ekat = str_tgf_format(str(value[5]))
                    # print(tgf_ekat)
                if value[6]:
                    tgf_omsk = str_tgf_format(str(value[6]))
                    # print(tgf_omsk)
                if value[7]:
                    tgf_novo = str_tgf_format(str(value[7]))
                    # print(tgf_novo)
                if value[8]:
                    tgf_tomsk = str_tgf_format(str(value[8]))
                    # print(tgf_tomsk)
                if value[9]:
                    tgf_more = str_tgf_format(str(value[9]))
                    # print(tgf_more)
                if value[10]:
                    tgf_tmn = str_tgf_format(str(value[10]))
                    # print(tgf_tmn)
                if value[11]:
                    tgf_kurgan = str_tgf_format(str(value[11]))
                    # print(tgf_kurgan)
                if value[12]:
                    # tgf = str_tgf_format(str(value[12]))
                    tgf = str_get_rgf(
                        tgf_kurgan, tgf_tmn, tgf_more, tgf_tomsk, tgf_novo, tgf_omsk, tgf_ekat, tgf_kras, tgf_ynao,
                        tgf_hmao, rgf)
                    print(tgf)

                if value[13]:  # Отчет
                    str_tmp1 = str(value[13]).strip().replace("_x001E_", "")
                    report_name = str_cleanup(str_tmp1)
                    # print(report_name)
                #
                if value[14]:  # Авторы
                    author_name = str(value[14]).strip()
                    # print(author_name)
                    # author_tmp = str_clean(author_name)
                #     if len(author_tmp) > 2:
                #         result = re.match(r'^0-9', author_tmp)
                #         result2 = re.match(r'[^\D]', author_tmp)
                #         if not result or not result2:
                #             authors.append(author_tmp)
                #
                if value[15]:  # Год
                    year_str = str(value[15]).strip()
                    year_str = year_str.replace(',', ' ')
                    year_tmp = year_str.split()
                    if len(year_tmp):  # разбираем такое 2009   2009   2010
                        year_int = int(year_tmp[0])  # забираем первый год
                        # print(year_int)
                # if value[16]:  # Территория
                #     territory_name = str(value[16]).strip()
                #     # print(territory_name)
                if value[16]:  # Субъект РФ
                    subrf_name = str(value[16]).strip()
                    # print(subrf_name)

                if value[17]:  # Листы
                    list_name = str(value[17]).strip()
                    # print(list_name)

                if value[18]:  # № партии
                    part_name = str(value[18]).strip()
                    # print(part_name)
                if value[19]:  # Площадь
                    areaoil = str(value[19]).strip()
                    # print(areaoil)

                if value[20]:  # Месторождение
                    field = str(value[20]).strip()
                    # print(field)

                if value[21]:  # ЛУ
                    lu = str(value[21]).strip()
                    # print(lu)

                if value[22]:  # ПИ
                    pi_name = str(value[22]).strip()
                    # print(pi_name)

                if value[23]:  # Источник финансирования
                    fin_name = str(value[23]).strip()
                    # print(fin_name)

                if value[24]:  # Организация
                    org_name = str(value[24]).strip()
                    # print(org_name)

                if value[25]:  # Отчет ЗапСибНИИГГ
                    zsniigg_report = str(value[25]).strip()
                    # print(zsniigg_report)

                if value[26]:  # Информационный отчет
                    inf_report = str(value[26]).strip()
                    # print(inf_report)

                if value[27]:  # Вид работ
                    vid_rab = str(value[27]).strip()
                    # print(vid_rab)

                if value[28]:  # Комментарии
                    comments = str(value[28]).strip()
                    print(comments)

                stmt = insert(M_REPORT_TGF).values(
                    folder_root=folder_root,
                    folder_link=folder_link,
                    folder_short=folder_short,
                    folder_name=folder_name,
                    rgf=rgf,
                    tgf_hmao=tgf_hmao,
                    tgf_ynao=tgf_ynao,
                    tgf_kras=tgf_kras,
                    tgf_ekat=tgf_ekat,
                    tgf_omsk=tgf_omsk,
                    tgf_novo=tgf_novo,
                    tgf_tomsk=tgf_tomsk,
                    tgf_more=tgf_more,
                    tgf_tmn=tgf_tmn,
                    tgf_kurgan=tgf_kurgan,
                    tgf=tgf,
                    report_name=report_name,
                    author_name=author_name,
                    year_str=year_str,
                    year_int=year_int,
                    territory_name='',
                    subrf_name=subrf_name,
                    list_name=list_name,
                    part_name=part_name,
                    areaoil=areaoil,
                    field=field,
                    lu=lu,
                    pi_name=pi_name,
                    fin_name=fin_name,
                    org_name=org_name,
                    zsniigg_report=zsniigg_report,
                    inf_report=inf_report,
                    vid_rab=vid_rab,
                    comments=comments,
                )
                await session.execute(stmt)
                await session.commit()

        # authors_str = ",".join(authors)
        # authors2 = authors_str.split(",")
        # authors_tmp2 = sorted(set(authors2))
        # for author in authors_tmp2:
        #     if len(author) > 2:
        #         await AUTHOR_M(author_name=author).save()

    except Exception as e:
        str_err = "Exception occurred " + str(e)
        content = {"msg": f"ERROR!!! cnt is {cnt}", "err": str(e)}
        print(str_err)


async def report_all_objects():
    content = {"msg": "Fail"}
    try:
        # content = {"msg": "Success", "count": 111111}
        async with async_session_maker() as session:

            res = await session.scalars(
                select(M_REPORT_TGF)
                # .order_by(M_REPORT_TGF.name_ru)
            )
            _all = res.all()
            _cnt = len(_all)

            content = {"msg": "Success", "count": _cnt, "data": _all}
        # log.info("ngp load successfully")
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all reports from {M_REPORT_TGF.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_get_all_count():
    content = {"msg": f"error"}
    try:
        async with async_session_maker() as session:
            res = await session.scalar(select(func.count(M_REPORT_TGF.id)))
            content = {"msg": "Success", "count": res}
            return content
    except Exception as e:
        cont_err = f"fail. can't read ext from table ({M_REPORT_TGF.__tablename__})"
        content = {"msg": "error", "data": f"Exception occurred {str(e)} . {cont_err}"}
        print(content)
    finally:
        if session is not None:
            await session.close()
    return content


async def report_all_no_folder():
    content = {"msg": "Fail"}
    try:
        # content = {"msg": "Success", "count": 111111}
        async with async_session_maker() as session:
            #  НГП
            res = await session.scalars(
                select(M_REPORT_TGF)
                .where(M_REPORT_TGF.folder_root == '')
                # .order_by(M_REPORT_TGF.name_ru)
            )
            _all = res.all()
            _cnt = len(_all)
            content = {"msg": "Success", "count": _cnt, "data": _all}
        # log.info("ngp load successfully")
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all reports from {M_REPORT_TGF.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_all_rtgf(tgf_param: mapped_column):
    content = {"msg": "Fail"}
    try:
        # content = {"msg": "Success", "count": 111111}
        async with async_session_maker() as session:
            #  НГП
            res = await session.scalars(
                select(M_REPORT_TGF)
                .where(tgf_param != '')
                .order_by(tgf_param)
            )
            _all = res.all()
            _cnt = len(_all)
            content = {"msg": "Success", "count": _cnt, "data": _all}
        # log.info("ngp load successfully")
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all reports from {M_REPORT_TGF.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_all_by_year(year: str):
    content = {"msg": "Fail"}
    try:
        # content = {"msg": "Success", "count": 111111}
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_REPORT_TGF)
                .where(M_REPORT_TGF.year_str.ilike(year))
            )
            _all = res.all()
            _cnt = len(_all)
            content = {"msg": "Success", "count": _cnt, "data": _all}
        # log.info("ngp load successfully")
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all reports from {M_REPORT_TGF.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_all_rgf():
    content = await report_all_rtgf(M_REPORT_TGF.rgf)
    return content


async def report_all_tgf_hmao():
    content = await report_all_rtgf(M_REPORT_TGF.tgf_hmao)
    return content


async def report_all_tgf_ynao():
    content = await report_all_rtgf(M_REPORT_TGF.tgf_ynao)
    return content


async def report_all_tgf_kras():
    content = await report_all_rtgf(M_REPORT_TGF.tgf_kras)
    return content


async def report_all_tgf_ekat():
    content = await report_all_rtgf(M_REPORT_TGF.tgf_ekat)
    return content


async def report_all_tgf_omsk():
    content = await report_all_rtgf(M_REPORT_TGF.tgf_omsk)
    return content


async def report_all_tgf_novo():
    content = await report_all_rtgf(M_REPORT_TGF.tgf_novo)
    return content


async def report_all_tgf_tomsk():
    content = await report_all_rtgf(M_REPORT_TGF.tgf_tomsk)
    return content


async def report_all_tgf_more():
    content = await report_all_rtgf(M_REPORT_TGF.tgf_more)
    return content


async def report_all_tgf_tmn():
    content = await report_all_rtgf(M_REPORT_TGF.tgf_tmn)
    return content


async def report_all_tgf_kurgan():
    content = await report_all_rtgf(M_REPORT_TGF.tgf_kurgan)
    return content


async def report_all_tgf():
    content = await report_all_rtgf(M_REPORT_TGF.tgf)
    return content


async def report_all_year():
    content = await report_all_rtgf(M_REPORT_TGF.year_str)
    return content


async def report_get_all_by_year(year: str):
    content = await report_all_by_year(year)
    return content


async def report_report_by_author(author: str):
    content = {"msg": "Fail"}
    try:
        authors = []
        authors_tmp2 = []
        async with async_session_maker() as session:
            author_str = f"%{author}%"
            res = await session.scalars(
                select(M_REPORT_TGF)
                .where(M_REPORT_TGF.author_name.ilike(author_str))
            )
            _all = res.all()
            _cnt = len(_all)
            content = {"msg": "Success", "count": _cnt, "data": _all}
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all reports from {M_REPORT_TGF.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_all_by_rgf(rgf: str):
    content = {"msg": "Fail"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_REPORT_TGF)
                .where(M_REPORT_TGF.rgf.ilike(rgf))
            )
            _all = res.all()
            _cnt = len(_all)
            content = {"msg": "Success", "count": _cnt, "data": _all}
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all reports from {M_REPORT_TGF.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_all_by_tgf_hmao(tgf_hmao: str):
    content = {"msg": "Fail"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_REPORT_TGF)
                .where(M_REPORT_TGF.tgf_hmao.ilike(tgf_hmao))
            )
            _all = res.all()
            _cnt = len(_all)
            content = {"msg": "Success", "count": _cnt, "data": _all}
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all reports from {M_REPORT_TGF.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_all_by_tgf_ynao(tgf_ynao: str):
    content = {"msg": "Fail"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_REPORT_TGF)
                .where(M_REPORT_TGF.tgf_ynao.ilike(tgf_ynao))
            )
            _all = res.all()
            _cnt = len(_all)
            content = {"msg": "Success", "count": _cnt, "data": _all}
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all reports from {M_REPORT_TGF.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_all_by_tgf_kras(tgf_kras: str):
    content = {"msg": "Fail"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_REPORT_TGF)
                .where(M_REPORT_TGF.tgf_kras.ilike(tgf_kras))
            )
            _all = res.all()
            _cnt = len(_all)
            content = {"msg": "Success", "count": _cnt, "data": _all}
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all reports from {M_REPORT_TGF.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_all_by_tgf_ekat(tgf_ekat: str):
    content = {"msg": "Fail"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_REPORT_TGF)
                .where(M_REPORT_TGF.tgf_ekat.ilike(tgf_ekat))
            )
            _all = res.all()
            _cnt = len(_all)
            content = {"msg": "Success", "count": _cnt, "data": _all}
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all reports from {M_REPORT_TGF.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_all_by_tgf_omsk(tgf_omsk: str):
    content = {"msg": "Fail"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_REPORT_TGF)
                .where(M_REPORT_TGF.tgf_omsk.ilike(tgf_omsk))
            )
            _all = res.all()
            _cnt = len(_all)
            content = {"msg": "Success", "count": _cnt, "data": _all}
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all reports from {M_REPORT_TGF.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_all_by_tgf_novo(tgf_novo: str):
    content = {"msg": "Fail"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_REPORT_TGF)
                .where(M_REPORT_TGF.tgf_novo.ilike(tgf_novo))
            )
            _all = res.all()
            _cnt = len(_all)
            content = {"msg": "Success", "count": _cnt, "data": _all}
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all reports from {M_REPORT_TGF.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_all_by_tgf_tomsk(tgf_tomsk: str):
    content = {"msg": "Fail"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_REPORT_TGF)
                .where(M_REPORT_TGF.tgf_tomsk.ilike(tgf_tomsk))
            )
            _all = res.all()
            _cnt = len(_all)
            content = {"msg": "Success", "count": _cnt, "data": _all}
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all reports from {M_REPORT_TGF.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_all_by_tgf_more(tgf_more: str):
    content = {"msg": "Fail"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_REPORT_TGF)
                .where(M_REPORT_TGF.tgf_more.ilike(tgf_more))
            )
            _all = res.all()
            _cnt = len(_all)
            content = {"msg": "Success", "count": _cnt, "data": _all}
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all reports from {M_REPORT_TGF.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_all_by_tgf_tmn(tgf_tmn: str):
    content = {"msg": "Fail"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_REPORT_TGF)
                .where(M_REPORT_TGF.tgf_tmn.ilike(tgf_tmn))
            )
            _all = res.all()
            _cnt = len(_all)
            content = {"msg": "Success", "count": _cnt, "data": _all}
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all reports from {M_REPORT_TGF.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_all_by_tgf_kurgan(tgf_kurgan: str):
    content = {"msg": "Fail"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_REPORT_TGF)
                .where(M_REPORT_TGF.tgf_kurgan.ilike(tgf_kurgan))
            )
            _all = res.all()
            _cnt = len(_all)
            content = {"msg": "Success", "count": _cnt, "data": _all}
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all reports from {M_REPORT_TGF.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_get_author():
    content = {"msg": "Fail"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_R_AUTHOR)
                .order_by(M_R_AUTHOR.name_ru)
            )
            _all = res.all()
            _cnt = len(_all)
            content = {"msg": "Success", "count": _cnt, "data": _all}
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all from {M_R_AUTHOR.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_get_author_by_id(id: int):
    content = {"msg": "Fail"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_R_AUTHOR)
                .where(M_R_AUTHOR.id == id)
                .order_by(M_R_AUTHOR.name_ru)

            )
            _all = res.all()
            _cnt = len(_all)
            content = {"msg": "Success", "count": _cnt, "data": _all}
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all from {M_R_AUTHOR.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_get_list():
    content = {"msg": "Fail"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_R_LIST)
                .order_by(M_R_LIST.name_ru)
            )
            _all = res.all()
            _cnt = len(_all)
            content = {"msg": "Success", "count": _cnt, "data": _all}
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all from {M_R_LIST.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_get_subrf():
    content = {"msg": "Fail"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_R_SUBRF)
                .order_by(M_R_SUBRF.name_ru)
            )
            _all = res.all()
            _cnt = len(_all)
            content = {"msg": "Success", "count": _cnt, "data": _all}
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all from {M_R_SUBRF.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_get_org():
    content = {"msg": "Fail"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_R_ORG)
                .order_by(M_R_ORG.name_ru)
            )
            _all = res.all()
            _cnt = len(_all)
            content = {"msg": "Success", "count": _cnt, "data": _all}
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all from {M_R_ORG.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_get_area():
    content = {"msg": "Fail"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_R_AREA)
                .order_by(M_R_AREA.name_ru)
            )
            _all = res.all()
            _cnt = len(_all)
            content = {"msg": "Success", "count": _cnt, "data": _all}
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all from {M_R_AREA.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_get_field():
    content = {"msg": "Fail"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_R_FIELD)
                .order_by(M_R_FIELD.name_ru)
            )
            _all = res.all()
            _cnt = len(_all)
            content = {"msg": "Success", "count": _cnt, "data": _all}
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all from {M_R_FIELD.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_get_lu():
    content = {"msg": "Fail"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_R_LU)
                .order_by(M_R_LU.name_ru)
            )
            _all = res.all()
            _cnt = len(_all)
            content = {"msg": "Success", "count": _cnt, "data": _all}
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all from {M_R_LU.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_get_pi():
    content = {"msg": "Fail"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_R_PI)
                .order_by(M_R_PI.name_ru)
            )
            _all = res.all()
            _cnt = len(_all)
            content = {"msg": "Success", "count": _cnt, "data": _all}
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all from {M_R_PI.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_get_vid_rab():
    content = {"msg": "Fail"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_R_VID_RAB)
                .order_by(M_R_VID_RAB.name_ru)
            )
            _all = res.all()
            _cnt = len(_all)
            content = {"msg": "Success", "count": _cnt, "data": _all}
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all from {M_R_VID_RAB.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content


async def report_get_model_all_count(model_param):
    content = {"msg": "Fail"}
    try:
        async with async_session_maker() as session:
            res = await session.scalar(select(func.count(model_param.id)))
            content = {"msg": "Success", "count": res}
            return content
    except Exception as e:
        cont_err = f"fail. can't read ext from table ({model_param.__tablename__})"
        content = {"msg": "error", "data": f"Exception occurred {str(e)} . {cont_err}"}
        print(content)
    finally:
        if session is not None:
            await session.close()
    return content


async def report_get_model_author_count():
    content = await report_get_model_all_count(M_R_AUTHOR)
    return content


async def report_get_model_list_count():
    content = await report_get_model_all_count(M_R_LIST)
    return content


async def report_get_model_subrf_count():
    content = await report_get_model_all_count(M_R_SUBRF)
    return content


async def report_get_model_org_count():
    content = await report_get_model_all_count(M_R_ORG)
    return content


async def report_get_model_area_count():
    content = await report_get_model_all_count(M_R_AREA)
    return content


async def report_get_model_field_count():
    content = await report_get_model_all_count(M_R_FIELD)
    return content


async def report_get_model_lu_count():
    content = await report_get_model_all_count(M_R_LU)
    return content


async def report_get_model_pi_count():
    content = await report_get_model_all_count(M_R_PI)
    return content


async def report_get_model_vid_rab_count():
    content = await report_get_model_all_count(M_R_VID_RAB)
    return content


async def report_index_create():
    content = {"msg": "Fail"}

    try:
        async with async_session_maker() as session:
            print(f"Удаляем индекс {cfg.REPORT_FTS_INDEX}")
            stmt = text(f"DROP INDEX IF EXISTS {cfg.REPORT_FTS_INDEX};")
            res = await session.execute(stmt)
            print(res)

            print(f"Создаем TSVector ...")
            stmt = text(
                f"UPDATE {M_REPORT_TGF.__tablename__} SET {M_REPORT_TGF.report_fts.key} = TO_TSVECTOR('russian', "
                f"{M_REPORT_TGF.report_name.key} ||' '|| "
                f"{M_REPORT_TGF.org_name.key} ||' '|| "
                f"{M_REPORT_TGF.author_name.key} ||' '|| "
                f"{M_REPORT_TGF.folder_link.key} ||' '|| "
                f"{M_REPORT_TGF.list_name.key} ||' '|| "
                f"{M_REPORT_TGF.part_name.key} ||' '|| "
                f"{M_REPORT_TGF.org_name.key} ||' '|| "
                f"{M_REPORT_TGF.areaoil.key} ||' '|| "
                f"{M_REPORT_TGF.field.key} ||' '|| "
                f"{M_REPORT_TGF.comments.key} ||' '|| "
                f"{M_REPORT_TGF.lu.key} ||' '|| "
                f"{M_REPORT_TGF.pi_name.key} ||' '|| "
                f"{M_REPORT_TGF.subrf_name.key} ||' '|| "
                f"{M_REPORT_TGF.territory_name.key} ||' '|| "
                f"{M_REPORT_TGF.tgf.key} ||' '|| "
                f"{M_REPORT_TGF.fin_name.key}) "
                f";")
            res = await session.execute(stmt)
            print(res)

            print(f"Создаем индекс ...")
            stmt = text(
                f"CREATE INDEX IF NOT EXISTS {cfg.REPORT_FTS_INDEX} ON {M_REPORT_TGF.__tablename__} USING GIN ({M_REPORT_TGF.report_fts.key});")
            print(stmt)
            res = await session.execute(stmt)
            print(res)
            await session.commit()
        print("OK")
        content = {"msg": "Success", "count": 0}

    except Exception as e:
        content = {"msg": f"can't create index {cfg.REPORT_FTS_INDEX}"}
        print("Exception occurred " + str(e))
        return content
    finally:
        if session is not None:
            await session.close()
    return content


# https://xata.io/blog/postgres-full-text-search-engine

async def report_fulltext_search(search_str: str, client_host: str):
    content = {"msg": "Success"}
    report_update_history.delay(search_str, client_host)  # сохраняем историю

    str_query_local = ''
    str_arr = search_str.split(" ")
    print(f"len of str_arr {len(str_arr)}")

    if len(str_arr) > 1:
        str_query_local = search_str.strip().lower().replace(" ", "&")
    else:
        str_query_local = f"{search_str.strip().lower()}:*"

    # str_query_local = search_str.strip().lower().replace(" ", "&")
    try:
        async with async_session_maker() as session:
            print(str_query_local)
            stmt = text(
                f"SELECT * FROM {M_REPORT_TGF.__tablename__} WHERE report_tgf.report_fts @@ to_tsquery('russian', '{str_query_local}');")
            # res = await session.execute(stmt)
            print(stmt)

            res = await session.scalars(select(M_REPORT_TGF).from_statement(stmt))
            print(res)

            # res = await session.scalars(
            #     select(M_REPORT_TGF)
            #     .where(M_REPORT_TGF.report_fts.match(str_query_local, postgresql_regconfig='russian'))
            # )
            _all = res.all()
            cnt = len(_all)
            content = {"msg": "Success", "count": cnt, "data": _all}
        return content
    except Exception as e:
        content = {"msg": f"can't search in index {cfg.REPORT_FTS_INDEX}"}
        print("Exception occurred " + str(e))
        return content
    finally:
        if session is not None:
            await session.close()


# async def report_update_history(query_str: str):

    # try:
    #     async with async_session_maker() as session:
    #         stmt = insert(M_HISTORY).values(
    #             search_str=query_str,
    #         )
    #         await session.execute(stmt)
    #         await session.commit()
    # except Exception as e:
    #     str_err = "Exception occurred " + str(e)
    #     content = {"msg": f"ERROR in report_update_history!!!", "err": str(e)}
    #     print(str_err)



# async def report_message_create(fio:str, email:str, message:str):
async def report_message_create(message: S_R_MESSAGE_POST):
    content = {"msg": "Success"}
    try:
        async with async_session_maker() as session:
            stmt = insert(M_R_MESSAGE).values(
                fio=message.fio,
                email=message.email,
                name_ru=message.name_ru,
            )
            res = await session.execute(stmt)
            await session.commit()
            _all = res.all()
            cnt = len(_all)
            content = {"msg": "Success", "count": cnt, "data": _all}
        return content
    except Exception as e:
        content = {"msg": f"can't insert into  {M_R_MESSAGE.__tablename__}"}
        print("Exception occurred " + str(e))
        return content
    finally:
        if session is not None:
            await session.close()


async def report_get_message():
    content = {"msg": "Fail"}
    try:
        async with async_session_maker() as session:
            res = await session.scalars(
                select(M_R_MESSAGE)
                .order_by(M_R_MESSAGE.id)
            )
            _all = res.all()
            _cnt = len(_all)
            content = {"msg": "Success", "count": _cnt, "data": _all}
        return content
    except Exception as e:
        content = {"msg": "Fail", "data": f"Can't get all from {M_R_MESSAGE.__tablename__}... "}
        print("Exception occurred " + str(e))
        # fastapi_logger.exception("update_user_password")
        return content
